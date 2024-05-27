import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def week6and7(file_path):
    # Load the workbook and select the worksheet
    wb = load_workbook(file_path)
    ws = wb['week6n7']
    
    # Find the last used column in the worksheet
    last_column = ws.max_column
    
    # Check if there are at least eight columns to remove
    if last_column >= 8:
        # Loop through the last eight columns and delete them
        for col in range(last_column, last_column - 8, -1):
            ws.delete_cols(col)
        
        # Find the last row and last column of data in the worksheet
        last_row = ws.max_row
        last_col = ws.max_column
        
        # Convert all headers to strings
        for col in range(1, last_col + 1):
            cell = ws.cell(row=1, column=col)
            if not isinstance(cell.value, str):
                cell.value = str(cell.value)
        
        # Define the data range dynamically based on the last row and last column
        data_range = f"A1:{get_column_letter(last_col)}{last_row}"
        
        # Define a table name (you can customize this)
        table_name = "MyTable"
        
        # Create the table
        table = Table(displayName=table_name, ref=data_range)
        
        # Add a table style (optional)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        
        # Add the table to the worksheet
        ws.add_table(table)

        # Define the column containing the school names (6th column)
        school_column = [row[0] for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True)]
        
        # Create a collection to store unique school names
        unique_schools = list(set(school_column))
        
        # Get the directory path from the full file path
        directory_path = os.path.dirname(file_path)
        
        # Loop through unique schools
        for school in unique_schools:
            if school is None:
                continue
            folder_path = os.path.join(directory_path, school)
            
            # Check if the folder already exists with the school name
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)  # Create a folder for the school if it doesn't exist
            
            # Create a new workbook and add the data worksheet
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = 'Data'
            
            # Filter and copy the header row
            header = [cell.value for cell in ws[1]]
            new_ws.append(header)
            
            # Loop through rows in the original sheet and copy rows matching the current school
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[5] == school:
                    new_ws.append(row)
            
            # Sanitize the school name to remove spaces for the table name
            sanitized_school_name = school.replace(' ', '_')
            
            # Convert the data into a table
            last_row = new_ws.max_row
            last_column = new_ws.max_column
            tbl_range = f"A1:{get_column_letter(last_column)}{last_row}"
            table = Table(displayName=f"SchoolTable_{sanitized_school_name}", ref=tbl_range)
            
            # Add a table style
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            new_ws.add_table(table)
            
            # Copy the original 'week1' sheet to the new workbook
            orig_week6and7_ws = wb['week6n7']
            new_week6and7_ws = new_wb.create_sheet(title='week6n7')
            
            for row in orig_week6and7_ws.iter_rows(values_only=True):
                new_week6and7_ws.append(row)
            
            # Convert the 'week1' sheet data into a table
            last_row_week6and7 = new_week6and7_ws.max_row
            last_column_week6and7 = new_week6and7_ws.max_column
            tbl_range_week6and7 = f"A1:{get_column_letter(last_column_week6and7)}{last_row_week6and7}"
            table_week6and7 = Table(displayName="Week9Table", ref=tbl_range_week6and7)
            
            # Add a table style for the 'week1' sheet
            style_week6and7 = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table_week6and7.tableStyleInfo = style_week6and7
            new_week6and7_ws.add_table(table_week6and7)
            
            # Save the new workbook in the school's folder with the name "week1.xlsx"
            new_file_path = os.path.join(folder_path, 'week6and7.xlsx')
            new_wb.save(new_file_path)
    else:
        print("There are not enough columns to remove.")

