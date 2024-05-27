import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from Chapter1.week1 import week1
from Chapter1.week2 import week2
from Chapter1.week3 import week3
from Chapter1.week4 import week4
from Chapter2.week5 import week5
from Chapter2.week6and7 import week6and7
from Chapter2.week8 import week8
from Chapter2.week9 import week9

selected_file_path = None  # Declare selected_file_path as a global variable
# Declare Chapter1window as a global variable
Chapter1window = None
Chapter2window = None

def browse_clicked():
    global selected_file_path  # Access the global variable
    # Create a file dialog
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=(("Excel Files", "*.xlsx; *.xlsm; *.xls"),)
    )

    # Check if a file was selected
    if file_path:
        # Assign the selected file path to the global variable
        selected_file_path = file_path
        # Clear the existing text in the textbox
        textbox.delete(1.0, tk.END)
        # Insert the selected file path into the textbox
        textbox.insert(tk.END, selected_file_path)

def Chapter1_clicked():
    Chapter1_window()

def Chapter1_window():
    global Chapter1window
    # Create a new window
    root.withdraw()  # Hide the main window
    Chapter1window= tk.Toplevel()
    Chapter1window.title("Chapter 1 Data Cleaning")
    Chapter1window.geometry("500x370")

    def on_closing():
        Chapter1window.destroy()  # Destroy the second window
        root.deiconify()  # Show the main window again

    # Set the fixed width and height for buttons
    button_width = 15
    button_height = 3

    # Create buttons
    button_week1 = tk.Button(Chapter1window, text="Week 1", width=button_width, height=button_height, command=week1_clicked)
    button_week2 = tk.Button(Chapter1window, text="Week 2", width=button_width, height=button_height, command=week2_clicked)
    button_week3 = tk.Button(Chapter1window, text="Week 3", width=button_width, height=button_height, command=week3_clicked)
    button_week4 = tk.Button(Chapter1window, text="Week 4", width=button_width, height=button_height, command=week4_clicked)
    button = tk.Button(Chapter1window, width=10, height=2, text="Exit ", command=on_closing)

    # Place buttons in the window
    button_week1.pack(pady=10)
    button_week2.pack(pady=10)
    button_week3.pack(pady=10)
    button_week4.pack(pady=10)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week1_clicked():
    # Call the function to open the Week 1 window
    open_week1_window()

def open_week1_window():
    global Chapter1window
    # # Hide the main window
    Chapter1window.withdraw()

    # Create a new window
    week1_window = tk.Toplevel()
    week1_window.title("Week 1 Window")
    week1_window.geometry("500x220")

    def on_closing():
        week1_window.destroy()  # Destroy the second window
        Chapter1window.deiconify()  # Show the main window again


    # Create a label and entry widget for the textbox
    label = tk.Label(week1_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week1_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week1_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week1_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week1_create)
    button_create.pack(pady=5)

    button = tk.Button(week1_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week1_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week1(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

def week2_clicked():
    # Call the function to open the Week 1 window
    open_week2_window()

def open_week2_window():
    global Chapter1window
    # # Hide the main window
    Chapter1window.withdraw()

    # Create a new window
    week2_window = tk.Toplevel()
    week2_window.title("Week 2 Window")
    week2_window.geometry("500x220")

    def on_closing():
        week2_window.destroy()  # Destroy the second window
        Chapter1window.deiconify()  # Show the main window again

    # Create a label and entry widget for the textbox
    label = tk.Label(week2_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week2_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week2_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week2_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week2_create)
    button_create.pack(pady=5)

    button = tk.Button(week2_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week2_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week2(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

def week3_clicked():
    # Call the function to open the Week 1 window
    open_week3_window()

def open_week3_window():
    global Chapter1window
    # # Hide the main window
    Chapter1window.withdraw()
    # Create a new window
    week3_window = tk.Toplevel()
    week3_window.title("Week 3 Window")
    week3_window.geometry("500x220")

    def on_closing():
        week3_window.destroy()  # Destroy the second window
        Chapter1window.deiconify()  # Show the main window again

    # Create a label and entry widget for the textbox
    label = tk.Label(week3_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week3_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week3_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week3_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week3_create)
    button_create.pack(pady=5)

    button = tk.Button(week3_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week3_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week3(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

def week4_clicked():
    # Call the function to open the Week 1 window
    open_week4_window()

def open_week4_window():
    global Chapter1window
    # # Hide the main window
    Chapter1window.withdraw()

    # Create a new window
    week4_window = tk.Toplevel()
    week4_window.title("Week 4 Window")
    week4_window.geometry("500x200")

    def on_closing():
        week4_window.destroy()  # Destroy the second window
        Chapter1window.deiconify()  # Show the main window again

    # Create a label and entry widget for the textbox
    label = tk.Label(week4_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week4_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week4_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week4_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command= week4_create)
    button_create.pack(pady=5)

    button = tk.Button(week4_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week4_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week4(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")


def Chapter2_clicked():
    Chapter2_window()

def Chapter2_window():
    global Chapter2window
     # Create a new window
    root.withdraw()  # Hide the main window
    # Create a new window
    Chapter2window= tk.Toplevel()
    Chapter2window.title("Chapter 2 Data Cleaning")
    Chapter2window.geometry("500x370")

    def on_closing():
        Chapter2window.destroy()  # Destroy the second window
        root.deiconify()  # Show the main window again

    # Set the fixed width and height for buttons
    button_width = 15
    button_height = 3

    # Create buttons
    button_week5 = tk.Button(Chapter2window, text="Week 5", width=button_width, height=button_height, command=week5_clicked)
    button_week6and7 = tk.Button(Chapter2window, text="Week 6 And 7", width=button_width, height=button_height, command=week6and7_clicked)
    button_week8 = tk.Button(Chapter2window, text="Week 8", width=button_width, height=button_height, command=week8_clicked)
    button_week9 = tk.Button(Chapter2window, text="Week 9", width=button_width, height=button_height, command=week9_clicked)
    button = tk.Button(Chapter2window, width=10, height=2, text="Exit ", command=on_closing)

    # Place buttons in the window
    button_week5.pack(pady=10)
    button_week6and7.pack(pady=10)
    button_week8.pack(pady=10)
    button_week9.pack(padx=10)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week5_clicked():
    # Call the function to open the Week 1 window
    open_week5_window()

def open_week5_window():
    global Chapter2window
    # # Hide the main window
    Chapter2window.withdraw()

    # Create a new window
    week5_window = tk.Toplevel()
    week5_window.title("Week 5 Window")
    week5_window.geometry("500x220")

    def on_closing():
        week5_window.destroy()  # Destroy the second window
        Chapter2window.deiconify()  # Show the main window again

    # Create a label and entry widget for the textbox
    label = tk.Label(week5_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week5_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week5_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week5_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week5_create)
    button_create.pack(pady=5)

    button = tk.Button(week5_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week5_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week5(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

def week6and7_clicked():
    # Call the function to open the Week 1 window
    open_week6and7_window()

def open_week6and7_window():
    global Chapter2window
    # # Hide the main window
    Chapter2window.withdraw()

    # Create a new window
    week6and7_window = tk.Toplevel()
    week6and7_window.title("Week 6and7 Window")
    week6and7_window.geometry("500x220")

    def on_closing():
        week6and7_window.destroy()  # Destroy the second window
        Chapter2window.deiconify()  # Show the main window again

    # Create a label and entry widget for the textbox
    label = tk.Label(week6and7_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week6and7_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week6and7_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week6and7_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week6and7_create)
    button_create.pack(pady=5)

    button = tk.Button(week6and7_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week6and7_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week6and7(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

def week8_clicked():
    # Call the function to open the Week 1 window
    open_week8_window()

def open_week8_window():
    global Chapter2window
    # # Hide the main window
    Chapter2window.withdraw()

    # Create a new window
    week8_window = tk.Toplevel()
    week8_window.title("Week 8 Window")
    week8_window.geometry("500x220")

    def on_closing():
        week8_window.destroy()  # Destroy the second window
        Chapter2window.deiconify()  # Show the main window again


    # Create a label and entry widget for the textbox
    label = tk.Label(week8_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week8_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week8_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week8_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week8_create)
    button_create.pack(pady=5)

    button = tk.Button(week8_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week8_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week8(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

def week9_clicked():
    # Call the function to open the Week 1 window
    open_week9_window()

def open_week9_window():
    global Chapter2window
    # # Hide the main window
    Chapter2window.withdraw()

    # Create a new window
    week9_window = tk.Toplevel()
    week9_window.title("Week 9 Window")
    week9_window.geometry("500x220")

    def on_closing():
        week9_window.destroy()  # Destroy the second window
        Chapter2window.deiconify()  # Show the main window again

    # Create a label and entry widget for the textbox
    label = tk.Label(week9_window, text="Select Excel File:")
    label.pack()

    global textbox
    textbox = tk.Text(week9_window, height=2, width=40)
    textbox.pack()

    button_width = 40
    button_height = 2

    # Create a "Browse" button
    browse_button = tk.Button(week9_window, text="Browse", width=10, height=2, command=browse_clicked)
    browse_button.pack(pady=5)

    button_create = tk.Button(week9_window, text="Creating School Folders and Dataset", width=button_width, height=button_height, command=week9_create)
    button_create.pack(pady=5)

    button = tk.Button(week9_window, width=10, height=2, text="Exit ", command=on_closing)
    button.pack(side=tk.BOTTOM, anchor='se', pady=10, padx=10)

def week9_create():
    selected_file_path = textbox.get("1.0", "end-1c").strip()

    if selected_file_path:
        wb = load_workbook(selected_file_path)
        # Assuming these functions are implemented
        week9(selected_file_path)

        # Close the workbook without saving changes
        wb.close()

        messagebox.showinfo("Success", "Tasks executed Successfully.")
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")

# Create the main window
root = tk.Tk()
root.title("Data Cleaning")

# Set window size to 800x800 pixels
root.geometry("400x200")

# Set the fixed width and height for buttons
button_width = 20
button_height = 3

# Create buttons
button_week1 = tk.Button(root, text="Chapter 1 Data Cleaning", width=button_width, height=button_height, command=Chapter1_clicked)
button_week2 = tk.Button(root, text="Chapter 2 Data Cleaning", width=button_width, height=button_height, command=Chapter2_clicked)

# Place buttons in the window
button_week1.pack(pady=10)
button_week2.pack(pady=10)


# Start the main event loop
root.mainloop()